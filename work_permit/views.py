import json
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.views.decorators.http import require_POST
from django.conf import settings

from .models import WorkPermit, PermitComment, PermitExtension, PermitApprovalStage, PermitRenewal, CHECKLIST_TEMPLATES, PERMIT_TYPE_CHOICES, get_wp_workflow_stages, get_wp_renewal_workflow_stages
from .forms import WorkPermitForm, PermitApprovalForm, PermitCloseForm, PermitCommentForm, PermitExtensionForm
from accounts.models import Employee


def _can_approve(user):
    return (
        user.is_superuser
        or getattr(user, 'perm_wp_approve', False)
        or user.role in ('administrator', 'management', 'president_plant_head', 'department_hod')
        or any(r in ('administrator', 'management', 'president_plant_head', 'department_hod')
               for r in user.get_additional_roles())
    )


# ── WP Workflow Helpers ───────────────────────────────────────────────────────

def _base_url(request=None):
    site_url = getattr(settings, 'SITE_BASE_URL', None)
    if site_url:
        return site_url.rstrip('/')
    return f"{request.scheme}://{request.get_host()}" if request else 'http://localhost:8000'


def _find_wp_approver(role, dept=None, designation=None):
    """Find the best matching employee for a WP workflow stage."""
    qs = Employee.objects.filter(is_active=True)

    # Safety Officer (HSEF): dept=HSEF AND (HOD role OR Safety Officer designation)
    if role == 'hsef_safety':
        hsef_qs = qs.filter(
            Q(department='HSEF') | Q(additional_departments__contains='|HSEF|')
        )
        # Prefer Safety Officer designation
        if designation:
            so = hsef_qs.filter(designation__icontains=designation).first()
            if so:
                return so
        # Fall back to HSEF HOD
        return hsef_qs.filter(
            Q(role='department_hod') | Q(additional_roles__contains='|department_hod|')
        ).first()

    role_q = Q(role=role) | Q(additional_roles__contains=f'|{role}|')
    qs = qs.filter(role_q)
    if dept:
        qs = qs.filter(
            Q(department=dept) | Q(additional_departments__contains=f'|{dept}|')
        )
    if designation:
        qs = qs.filter(designation__icontains=designation)
    return qs.first()


def _create_wp_stages(permit):
    """Create PermitApprovalStage records based on isolation/moc flags."""
    permit.approval_stages.filter(workflow_type='main').delete()
    stages = get_wp_workflow_stages(permit.isolation_required, permit.moc_required)
    for i, (label, role, dept, desig, stype) in enumerate(stages):
        PermitApprovalStage.objects.create(
            permit=permit, stage=i, stage_label=label,
            approver_role=role, approver_dept=dept or '',
            approver_desig=desig or '', stage_type=stype, workflow_type='main',
        )


def _create_wp_renewal_stages(permit, renewal):
    """Create renewal workflow stages (acknowledge → approval → acknowledge)."""
    stages = get_wp_renewal_workflow_stages()
    for i, (label, role, dept, desig, stype) in enumerate(stages):
        PermitApprovalStage.objects.create(
            permit=permit, stage=i, stage_label=label,
            approver_role=role, approver_dept=dept or '',
            approver_desig=desig or '', stage_type=stype,
            workflow_type='renewal', renewal=renewal,
        )

def _main_workflow_stages(permit):
    return permit.approval_stages.filter(workflow_type='main').order_by('stage')


def _renewal_workflow_stages(renewal):
    return renewal.approval_stages.order_by('stage')


def _active_renewal(permit):
    return permit.renewals.filter(status='pending').order_by('-renewal_no').first()


def _current_pending_stage(permit):
    active_renewal = _active_renewal(permit)
    if active_renewal:
        return _renewal_workflow_stages(active_renewal).filter(status='pending').first()
    return _main_workflow_stages(permit).filter(status='pending').first()


def _send_wp_stage_email(permit, stage_obj, request=None):
    """Send approval-request or acknowledge-only email to the stage approver."""
    try:
        approver = _find_wp_approver(
            stage_obj.approver_role,
            dept=stage_obj.approver_dept or None,
            designation=stage_obj.approver_desig or None,
        )
        if not approver or not approver.email:
            return
        base = _base_url(request)
        is_ack = stage_obj.stage_type == 'acknowledge'
        all_stages = list(_renewal_workflow_stages(stage_obj.renewal) if stage_obj.workflow_type == 'renewal' else _main_workflow_stages(permit))
        total = len(all_stages)
        subject = f'[{"Acknowledgement" if is_ack else "Action Required"}] WP {"Renewal " if permit.renewal_required else ""}— Stage {stage_obj.stage + 1}/{total}: {permit.permit_number}'

        plain = (
            f"Dear {approver.employee_name},\n\n"
            f"Work Permit {permit.permit_number} {'renewal ' if permit.renewal_required else ''}requires your {'acknowledgement' if is_ack else 'approval'} at Stage {stage_obj.stage + 1}: {stage_obj.stage_label}.\n\n"
            f"Title: {permit.title}\nLocation: {permit.location}\nRisk: {permit.get_risk_level_display()}\n\n"
        )
        if is_ack:
            ack_url = f"{base}/work-permit/stage/{stage_obj.token}/approve/"
            plain += f"ACKNOWLEDGE: {ack_url}\n\nRegards,\nERP Department - Unity Cement"
        else:
            approve_url = f"{base}/work-permit/stage/{stage_obj.token}/approve/"
            reject_url  = f"{base}/work-permit/stage/{stage_obj.token}/reject/"
            plain += f"APPROVE: {approve_url}\nREJECT : {reject_url}\n\nRegards,\nERP Department - Unity Cement"

        progress_cells = []
        for idx, s in enumerate(all_stages):
            is_cur = s.pk == stage_obj.pk
            if s.status == 'approved':   bg, icon, fg = '#d1fae5', '&#10003;', '#065f46'
            elif s.status == 'rejected': bg, icon, fg = '#fee2e2', '&#10007;', '#991b1b'
            elif is_cur:                 bg, icon, fg = '#fef3c7', '&#9654;', '#92400e'
            else:                        bg, icon, fg = '#f1f5f9', '&#9679;', '#94a3b8'
            border = 'border:2px solid #2563eb;' if is_cur else ''
            short = s.stage_label.split('\u2014')[0].strip()[:12]
            progress_cells.append(
                f'<td style="text-align:center;padding:0 4px;">'
                f'<div style="display:inline-flex;flex-direction:column;align-items:center;gap:4px;">'
                f'<div style="width:34px;height:34px;border-radius:50%;background:{bg};color:{fg};'
                f'display:flex;align-items:center;justify-content:center;font-size:14px;font-weight:700;{border}">{icon}</div>'
                f'<div style="font-size:9px;font-weight:600;color:{fg};white-space:nowrap;">{short}</div>'
                f'</div></td>'
            )
            if idx < len(all_stages) - 1:
                conn = '#10b981' if s.status == 'approved' else '#e2e8f0'
                progress_cells.append(f'<td style="width:16px;"><div style="height:2px;background:{conn};margin-bottom:20px;"></div></td>')
        progress_html = ''.join(progress_cells)

        html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f0f4f8;font-family:Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f4f8;padding:28px 0;">
<tr><td align="center">
<table width="600" cellpadding="0" cellspacing="0" style="background:white;border-radius:16px;overflow:hidden;">
  <tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22);height:4px;"></td></tr>
  <tr><td style="background:linear-gradient(135deg,#0f172a,#1e3a5f,#1e40af);padding:32px 36px;text-align:center;">
    <div style="font-size:10px;color:rgba(255,255,255,0.5);letter-spacing:3px;text-transform:uppercase;margin-bottom:8px;">Unity Cement &mdash; ERP Department</div>
    <div style="font-size:24px;font-weight:800;color:white;margin-bottom:6px;">Work Permit {'Renewal ' if permit.renewal_required else ''}{'Acknowledgement' if is_ack else 'Approval'}</div>
    <div style="font-size:13px;color:rgba(255,255,255,0.7);margin-bottom:14px;">Stage {stage_obj.stage + 1} of {total}: <strong style="color:white;">{stage_obj.stage_label}</strong></div>
    <div style="display:inline-block;background:rgba(240,165,0,0.2);border:1.5px solid rgba(240,165,0,0.5);border-radius:50px;padding:5px 18px;font-size:13px;color:#fbbf24;font-weight:700;">{permit.permit_number}</div>
  </td></tr>
  <tr><td style="padding:20px 36px;background:#f8fafc;border-bottom:1px solid #e2e8f0;">
    <div style="font-size:10px;font-weight:700;color:#94a3b8;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:12px;">Approval Progress</div>
    <table cellpadding="0" cellspacing="0" style="width:100%;"><tr style="vertical-align:middle;">{progress_html}</tr></table>
  </td></tr>
  <tr><td style="padding:24px 36px 0;">
    <p style="font-size:15px;color:#1e293b;margin:0 0 6px;">Dear <strong style="color:#1e3a5f;">{approver.employee_name}</strong>,</p>
    <p style="font-size:13px;color:#64748b;margin:0;line-height:1.6;">Work Permit <strong>{permit.permit_number}</strong> requires your {'acknowledgement' if is_ack else 'approval'}. Please review and take action.</p>
  </td></tr>
  <tr><td style="padding:16px 36px;">
    <div style="border-radius:10px;overflow:hidden;border:1px solid #e2e8f0;">
      <div style="background:linear-gradient(135deg,#1e3a5f,#1e40af);padding:10px 14px;"><span style="font-size:10px;font-weight:700;color:white;letter-spacing:1.5px;text-transform:uppercase;">&#128203; Permit Details</span></div>
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr style="background:white;"><td style="padding:8px 14px;font-size:11px;font-weight:600;color:#475569;width:38%;border-bottom:1px solid #f1f5f9;">Permit No.</td><td style="padding:8px 14px;font-size:12px;color:#1e293b;border-bottom:1px solid #f1f5f9;"><strong>{permit.permit_number}</strong></td></tr>
        <tr style="background:#f8fafc;"><td style="padding:8px 14px;font-size:11px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Title</td><td style="padding:8px 14px;font-size:12px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{permit.title}</td></tr>
        <tr style="background:white;"><td style="padding:8px 14px;font-size:11px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Location</td><td style="padding:8px 14px;font-size:12px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{permit.location}</td></tr>
        <tr style="background:#f8fafc;"><td style="padding:8px 14px;font-size:11px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Permit Type</td><td style="padding:8px 14px;font-size:12px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{permit.get_permit_type_display()}</td></tr>
        <tr style="background:white;"><td style="padding:8px 14px;font-size:11px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Risk Level</td><td style="padding:8px 14px;font-size:12px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{permit.get_risk_level_display()}</td></tr>
        <tr style="background:#f8fafc;"><td style="padding:8px 14px;font-size:11px;font-weight:600;color:#475569;border-bottom:1px solid #f1f5f9;">Requested By</td><td style="padding:8px 14px;font-size:12px;color:#1e293b;border-bottom:1px solid #f1f5f9;">{permit.requested_by.employee_name} ({permit.requested_by.department})</td></tr>
        <tr style="background:white;"><td style="padding:8px 14px;font-size:11px;font-weight:600;color:#475569;">Isolation / MOC</td><td style="padding:8px 14px;font-size:12px;color:#1e293b;">Isolation: {'Yes' if permit.isolation_required else 'No'} &nbsp;|&nbsp; MOC: {'Yes' if permit.moc_required else 'No'}</td></tr>
      </table>
    </div>
  </td></tr>
  <tr><td style="padding:8px 36px 28px;">
    <div style="background:#f8fafc;border-radius:10px;padding:20px;border:1px solid #e2e8f0;">
      {'<p style="font-size:12px;color:#475569;font-weight:600;margin:0 0 14px;text-align:center;">&#9889; Take Action Now</p><table cellpadding="0" cellspacing="0" width="100%"><tr><td width="47%" align="center"><a href="' + approve_url + '" style="display:block;background:linear-gradient(135deg,#059669,#10b981);color:white;text-decoration:none;padding:14px 20px;border-radius:10px;font-size:15px;font-weight:800;text-align:center;">&#10003;&nbsp;&nbsp;APPROVE</a></td><td width="6%"></td><td width="47%" align="center"><a href="' + reject_url + '" style="display:block;background:linear-gradient(135deg,#dc2626,#ef4444);color:white;text-decoration:none;padding:14px 20px;border-radius:10px;font-size:15px;font-weight:800;text-align:center;">&#10007;&nbsp;&nbsp;REJECT</a></td></tr></table><p style="font-size:10px;color:#94a3b8;text-align:center;margin:12px 0 0;">Clicking opens a secure confirmation page where you can add remarks.</p>' if not is_ack else '<p style="font-size:12px;color:#475569;font-weight:600;margin:0 0 14px;text-align:center;">&#128276; Acknowledgement Required</p><table cellpadding="0" cellspacing="0" width="100%"><tr><td align="center"><a href="' + ack_url + '" style="display:block;background:linear-gradient(135deg,#2563eb,#1d4ed8);color:white;text-decoration:none;padding:14px 20px;border-radius:10px;font-size:15px;font-weight:800;text-align:center;">&#10003;&nbsp;&nbsp;ACKNOWLEDGE</a></td></tr></table><p style="font-size:10px;color:#94a3b8;text-align:center;margin:12px 0 0;">This is an acknowledgement only. No approval action is required.</p>'}
    </div>
  </td></tr>
  <tr><td style="background:linear-gradient(90deg,#f0a500,#e67e22);height:3px;"></td></tr>
  <tr><td style="background:#f8fafc;padding:16px 36px;text-align:center;">
    <p style="font-size:10px;color:#94a3b8;margin:0;">Automated email from <strong>ERP Department &mdash; Unity Cement</strong>. Do not reply.</p>
  </td></tr>
</table></td></tr></table></body></html>"""

        from django.core.mail import EmailMultiAlternatives
        msg = EmailMultiAlternatives(subject, plain, settings.DEFAULT_FROM_EMAIL, [approver.email])
        msg.attach_alternative(html, 'text/html')
        msg.send(fail_silently=True)
    except Exception:
        pass


def _send_wp_status_email(permit, action):
    """Notify the initiator about final approval or rejection."""
    try:
        if not permit.requested_by.email:
            return
        subject = f'Work Permit {action.upper()} — {permit.permit_number}'
        trail_rows = ''
        for i, s in enumerate(permit.approval_stages.all()):
            bg = 'white' if i % 2 == 0 else '#f8fafc'
            if s.status == 'approved':   bb, bf = '#d1fae5', '#065f46'
            elif s.status == 'rejected': bb, bf = '#fee2e2', '#991b1b'
            else:                        bb, bf = '#fef3c7', '#92400e'
            approver_name = s.approver.employee_name if s.approver else 'N/A'
            trail_rows += (
                f'<tr style="background:{bg};">'
                f'<td style="padding:7px 12px;font-size:11px;font-weight:600;color:#475569;">{s.stage + 1}</td>'
                f'<td style="padding:7px 12px;font-size:11px;color:#1e293b;">{s.stage_label}</td>'
                f'<td style="padding:7px 12px;font-size:11px;color:#475569;">{approver_name}</td>'
                f'<td style="padding:7px 12px;"><span style="font-size:10px;font-weight:700;padding:2px 8px;border-radius:20px;background:{bb};color:{bf};">{s.status.upper()}</span></td>'
                f'<td style="padding:7px 12px;font-size:10px;color:#64748b;">{s.remarks or "—"}</td>'
                '</tr>'
            )
        color = '#059669' if action == 'approved' else '#dc2626'
        plain = f"Dear {permit.requested_by.employee_name},\n\nYour Work Permit {permit.permit_number} has been {action}.\n\nRegards,\nERP Department - Unity Cement"
        html = f"""<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;background:#f0f4f8;padding:24px;">
<div style="max-width:600px;margin:0 auto;background:white;border-radius:14px;overflow:hidden;">
  <div style="background:{color};padding:28px 32px;text-align:center;">
    <div style="font-size:22px;font-weight:800;color:white;">Work Permit {action.title()}</div>
    <div style="margin-top:10px;display:inline-block;background:rgba(255,255,255,0.2);border-radius:50px;padding:4px 16px;font-size:13px;color:white;font-weight:600;">{permit.permit_number}</div>
  </div>
  <div style="padding:24px 32px;">
    <p style="font-size:15px;color:#1e293b;">Dear <strong>{permit.requested_by.employee_name}</strong>,</p>
    <p style="font-size:13px;color:#64748b;">Your Work Permit has been <strong>{action}</strong>.</p>
    <div style="border-radius:10px;overflow:hidden;border:1px solid #e2e8f0;margin-top:16px;">
      <div style="background:#f1f5f9;padding:10px 14px;border-bottom:1px solid #e2e8f0;"><span style="font-size:10px;font-weight:700;color:#475569;letter-spacing:1.5px;text-transform:uppercase;">Approval Trail</span></div>
      <table width="100%" cellpadding="0" cellspacing="0">
        <tr style="background:#f8fafc;"><th style="padding:7px 12px;font-size:10px;color:#94a3b8;text-align:left;">#</th><th style="padding:7px 12px;font-size:10px;color:#94a3b8;text-align:left;">Stage</th><th style="padding:7px 12px;font-size:10px;color:#94a3b8;text-align:left;">Approver</th><th style="padding:7px 12px;font-size:10px;color:#94a3b8;text-align:left;">Status</th><th style="padding:7px 12px;font-size:10px;color:#94a3b8;text-align:left;">Remarks</th></tr>
        {trail_rows}
      </table>
    </div>
  </div>
  <div style="background:#f8fafc;padding:14px 32px;text-align:center;"><p style="font-size:10px;color:#94a3b8;margin:0;">Automated email — ERP Department, Unity Cement</p></div>
</div></body></html>"""
        from django.core.mail import EmailMultiAlternatives
        msg = EmailMultiAlternatives(subject, plain, settings.DEFAULT_FROM_EMAIL, [permit.requested_by.email])
        msg.attach_alternative(html, 'text/html')
        msg.send(fail_silently=True)
    except Exception:
        pass


def _advance_wp_workflow(permit, request=None, renewal=None):
    """Move to next pending stage or mark fully approved.
    Acknowledge stages are auto-advanced after sending email (no action needed).
    """
    stage_qs = _renewal_workflow_stages(renewal) if renewal else _main_workflow_stages(permit)
    for stage in stage_qs:
        if stage.status == 'pending':
            if stage.stage_type == 'acknowledge':
                # Send ack email then auto-mark as approved and continue
                _send_wp_stage_email(permit, stage, request=request)
                approver = _find_wp_approver(
                    stage.approver_role,
                    dept=stage.approver_dept or None,
                    designation=stage.approver_desig or None,
                )
                stage.status = 'approved'
                stage.approver = approver
                stage.acted_at = timezone.now()
                stage.remarks = 'Acknowledged (auto)'
                stage.save()
                continue  # check next stage
            else:
                # Approval stage — send email and wait
                permit.status = 'pending'
                permit.save()
                _send_wp_stage_email(permit, stage, request=request)
                return
    # All stages done → approved
    permit.status = 'approved'
    if not permit.actual_start:
        permit.actual_start = timezone.now()
    last_stage = stage_qs.filter(status='approved').last()
    if not renewal:
        permit.final_approved_by = last_stage.approver if last_stage else None
        permit.final_approved_at = timezone.now()
    permit.save()
    # Mark renewal record as approved without touching the original workflow
    if renewal:
        from datetime import time as _time
        deadline_date = permit.actual_start.date() + timezone.timedelta(days=7)
        valid_until = timezone.datetime.combine(deadline_date, _time(23, 59, 59),
                                                tzinfo=permit.actual_start.tzinfo)
        renewal.status = 'approved'
        renewal.approved_at = timezone.now()
        renewal.valid_from = permit.actual_start
        renewal.valid_until = valid_until
        renewal.save()
    _send_wp_status_email(permit, 'approved')


def _can_delete(user):
    return user.is_superuser or user.role == 'administrator'


def _can_write(user):
    return getattr(user, 'perm_wp_write', True)


def _can_view(user):
    return getattr(user, 'perm_wp_view', True)


def _wp_pending_stage_for_user(user, permit):
    """Return the pending stage this user can act on, or None."""
    if user.is_superuser or user.role == 'administrator':
        return _current_pending_stage(permit)
    stage = _current_pending_stage(permit)
    if stage and _wp_user_can_act(user, stage):
        return stage
    return None


@login_required
def dashboard(request):
    if not _can_view(request.user):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:index')

    qs = WorkPermit.objects.all()
    if not (request.user.is_superuser or _can_approve(request.user)):
        qs = qs.filter(requested_by=request.user)

    # Auto-expire approved permits past renewal deadline (midnight after 7 days)
    expired_ids = [p.pk for p in qs.filter(status='approved') if p.is_expired]
    if expired_ids:
        WorkPermit.objects.filter(pk__in=expired_ids).update(status='closed')
        qs = WorkPermit.objects.all()
        if not (request.user.is_superuser or _can_approve(request.user)):
            qs = qs.filter(requested_by=request.user)

    stats = {
        'total':     qs.count(),
        'active':    qs.filter(status='approved').count(),
        'pending':   qs.filter(status='pending').count(),
        'expired':   qs.filter(status='expired').count(),
        'closed':    qs.filter(status='closed').count(),
        'draft':     qs.filter(status='draft').count(),
        'rejected':  qs.filter(status='rejected').count(),
        'suspended': qs.filter(status='suspended').count(),
    }

    by_type = {}
    for code, label in PERMIT_TYPE_CHOICES:
        cnt = qs.filter(permit_type=code).count()
        if cnt:
            by_type[code] = {'label': label, 'count': cnt}

    recent = qs.order_by('-created_at')[:10]
    active_permits = qs.filter(status='approved').order_by('end_datetime')[:5]

    return render(request, 'work_permit/dashboard.html', {
        'stats': stats,
        'by_type': by_type,
        'recent': recent,
        'active_permits': active_permits,
        'permit_type_choices': PERMIT_TYPE_CHOICES,
    })


@login_required
def permit_list(request):
    if not _can_view(request.user):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:index')

    qs = WorkPermit.objects.select_related('requested_by')
    if not (request.user.is_superuser or _can_approve(request.user)):
        qs = qs.filter(requested_by=request.user)

    # Filters
    status_f = request.GET.get('status', '')
    type_f   = request.GET.get('permit_type', '')
    risk_f   = request.GET.get('risk_level', '')
    q        = request.GET.get('q', '').strip()

    if status_f:
        qs = qs.filter(status=status_f)
    if type_f:
        qs = qs.filter(permit_type=type_f)
    if risk_f:
        qs = qs.filter(risk_level=risk_f)
    if q:
        qs = qs.filter(
            Q(permit_number__icontains=q) | Q(title__icontains=q) |
            Q(location__icontains=q) | Q(contractor_name__icontains=q)
        )

    # Bulk delete
    if request.method == 'POST' and request.POST.get('action') == 'bulk_delete':
        if _can_delete(request.user):
            ids = request.POST.getlist('selected_ids')
            if ids:
                deleted, _ = WorkPermit.objects.filter(pk__in=ids).delete()
                messages.success(request, f'{deleted} permit(s) deleted.')
        else:
            messages.error(request, 'Only administrators can delete permits.')
        return redirect(request.get_full_path())

    paginator = Paginator(qs, 20)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    return render(request, 'work_permit/permit_list.html', {
        'permits': page_obj,
        'page_obj': page_obj,
        'permit_type_choices': PERMIT_TYPE_CHOICES,
        'status_f': status_f, 'type_f': type_f, 'risk_f': risk_f, 'q': q,
        'can_delete': _can_delete(request.user),
    })


@login_required
def permit_create(request):
    if not _can_write(request.user):
        messages.error(request, 'Access denied.')
        return redirect('work_permit:list')

    if request.method == 'POST':
        form = WorkPermitForm(request.POST, request.FILES)
        if form.is_valid():
            permit = form.save(commit=False)
            permit.requested_by = request.user
            # Save checklist
            ptype = form.cleaned_data['permit_type']
            checklist = {}
            for item in CHECKLIST_TEMPLATES.get(ptype, []):
                checklist[item['key']] = request.POST.get(f'chk_{item["key"]}') == 'on'
            permit.checklist_data = checklist
            permit.save()
            messages.success(request, f'Work Permit {permit.permit_number} created successfully.')
            return redirect('work_permit:detail', pk=permit.pk)
    else:
        form = WorkPermitForm()

    return render(request, 'work_permit/permit_form.html', {
        'form': form,
        'checklist_templates': CHECKLIST_TEMPLATES,
        'is_edit': False,
    })


@login_required
def permit_edit(request, pk):
    permit = get_object_or_404(WorkPermit, pk=pk)
    if permit.status not in ('draft', 'rejected') and not request.user.is_superuser:
        messages.error(request, 'Only draft or rejected permits can be edited.')
        return redirect('work_permit:detail', pk=pk)
    if permit.requested_by != request.user and not request.user.is_superuser:
        messages.error(request, 'Access denied.')
        return redirect('work_permit:list')

    if request.method == 'POST':
        form = WorkPermitForm(request.POST, request.FILES, instance=permit)
        if form.is_valid():
            p = form.save(commit=False)
            ptype = form.cleaned_data['permit_type']
            checklist = {}
            for item in CHECKLIST_TEMPLATES.get(ptype, []):
                checklist[item['key']] = request.POST.get(f'chk_{item["key"]}') == 'on'
            p.checklist_data = checklist
            p.save()
            messages.success(request, 'Permit updated.')
            return redirect('work_permit:detail', pk=pk)
    else:
        form = WorkPermitForm(instance=permit)

    return render(request, 'work_permit/permit_form.html', {
        'form': form,
        'permit': permit,
        'checklist_templates': CHECKLIST_TEMPLATES,
        'is_edit': True,
    })


@login_required
def permit_detail(request, pk):
    permit = get_object_or_404(WorkPermit, pk=pk)
    if not (request.user.is_superuser or _can_approve(request.user) or permit.requested_by == request.user):
        messages.error(request, 'Access denied.')
        return redirect('work_permit:list')

    comment_form = PermitCommentForm()
    ext_form = PermitExtensionForm()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'comment':
            comment_form = PermitCommentForm(request.POST)
            if comment_form.is_valid():
                c = comment_form.save(commit=False)
                c.permit = permit
                c.author = request.user
                c.save()
                messages.success(request, 'Comment added.')
                return redirect('work_permit:detail', pk=pk)
        elif action == 'extend':
            ext_form = PermitExtensionForm(request.POST)
            if ext_form.is_valid():
                ext = ext_form.save(commit=False)
                ext.permit = permit
                ext.requested_by = request.user
                ext.save()
                messages.success(request, 'Extension request submitted.')
                return redirect('work_permit:detail', pk=pk)

    stages = _main_workflow_stages(permit)
    pending_stage = _wp_pending_stage_for_user(request.user, permit)
    active_renewal = _active_renewal(permit)

    return render(request, 'work_permit/permit_detail.html', {
        'permit': permit,
        'comment_form': comment_form,
        'ext_form': ext_form,
        'can_approve': _can_approve(request.user),
        'is_admin': request.user.is_superuser or request.user.role == 'administrator',
        'stages': stages,
        'pending_stage': pending_stage,
        'renewals': permit.renewals.all(),
        'active_renewal': active_renewal,
        'renewal_stages': _renewal_workflow_stages(active_renewal) if active_renewal else [],
    })


@login_required
def permit_delete(request, pk):
    if not _can_delete(request.user):
        messages.error(request, 'Only administrators can delete permits.')
        return redirect('work_permit:detail', pk=pk)
    permit = get_object_or_404(WorkPermit, pk=pk)
    if request.method == 'POST':
        num = permit.permit_number
        permit.delete()
        messages.success(request, f'Permit {num} deleted.')
        return redirect('work_permit:list')
    return render(request, 'work_permit/permit_delete_confirm.html', {'permit': permit})


@login_required
def permit_submit(request, pk):
    permit = get_object_or_404(WorkPermit, pk=pk, requested_by=request.user)
    if permit.status == 'draft':
        _create_wp_stages(permit)
        permit.status = 'pending'
        permit.save()
        _advance_wp_workflow(permit, request=request)
        messages.success(request, f'{permit.permit_number} submitted for approval.')
    return redirect('work_permit:detail', pk=pk)


@login_required
def permit_approve(request, pk):
    permit = get_object_or_404(WorkPermit, pk=pk)

    # Admin can act on any pending stage; others only on their assigned stage
    is_admin = request.user.is_superuser or request.user.role == 'administrator'
    pending_stage = _current_pending_stage(permit)
    if not pending_stage:
        # Allow suspend on active permits
        if request.method == 'POST' and request.POST.get('action') == 'suspend':
            permit.status = 'suspended'
            permit.suspension_reason = request.POST.get('remarks', '')
            permit.save()
            messages.warning(request, f'{permit.permit_number} suspended.')
        else:
            messages.info(request, 'No pending approval stage.')
        return redirect('work_permit:detail', pk=pk)

    # Check if current user is the expected approver (admin can act on any stage)
    if not is_admin and not _wp_user_can_act(request.user, pending_stage):
        messages.error(request, 'You are not the expected approver for the current stage.')
        return redirect('work_permit:detail', pk=pk)

    if request.method == 'POST':
        form = PermitApprovalForm(request.POST)
        if form.is_valid():
            action  = form.cleaned_data['action']
            remarks = form.cleaned_data['remarks']
            if action == 'approve':
                pending_stage.status   = 'approved'
                pending_stage.approver = request.user
                pending_stage.remarks  = remarks
                pending_stage.acted_at = timezone.now()
                pending_stage.save()
                # Set actual_start on first approval if not set
                if not permit.actual_start:
                    permit.actual_start = timezone.now()
                    permit.save()
                _advance_wp_workflow(
                    permit,
                    request=request,
                    renewal=pending_stage.renewal if pending_stage.workflow_type == 'renewal' else None,
                )
                messages.success(request, f'Stage {pending_stage.stage + 1} approved.')
            elif action == 'reject':
                pending_stage.status   = 'rejected'
                pending_stage.approver = request.user
                pending_stage.remarks  = remarks
                pending_stage.acted_at = timezone.now()
                pending_stage.save()
                permit.status = 'rejected'
                permit.rejection_reason = remarks
                permit.save()
                _send_wp_status_email(permit, 'rejected')
                messages.warning(request, f'{permit.permit_number} rejected.')
            elif action == 'suspend':
                permit.status = 'suspended'
                permit.suspension_reason = remarks
                permit.save()
                messages.warning(request, f'{permit.permit_number} suspended.')
            return redirect('work_permit:detail', pk=pk)
    else:
        form = PermitApprovalForm()

    return render(request, 'work_permit/permit_approve.html', {
        'permit': permit,
        'form': form,
        'stage': pending_stage,
        'stages': _renewal_workflow_stages(pending_stage.renewal) if pending_stage and pending_stage.workflow_type == 'renewal' else _main_workflow_stages(permit),
        'is_admin': is_admin,
    })


def _wp_user_can_act(user, stage):
    """Check if user matches the expected approver for a WP stage."""
    role = stage.approver_role

    # Safety Officer (HSEF): dept=HSEF AND (role=department_hod OR designation contains 'Safety Officer')
    if role == 'hsef_safety':
        if not user.has_department('HSEF'):
            return False
        return user.has_role('department_hod') or (
            stage.approver_desig and stage.approver_desig.lower() in user.designation.lower()
        )

    # Standard role check
    if not user.has_role(role):
        return False
    if stage.approver_dept and not user.has_department(stage.approver_dept):
        return False
    if stage.approver_desig and stage.approver_desig.lower() not in user.designation.lower():
        return False
    return True


def wp_stage_action(request, token, action):
    """Token-based approval from email link — shows Approve & Reject buttons."""
    stage_obj = get_object_or_404(PermitApprovalStage, token=token)
    permit = stage_obj.permit

    if stage_obj.status != 'pending':
        return render(request, 'work_permit/wp_email_action_done.html', {
            'message': f'This stage has already been {stage_obj.status}.',
            'permit': permit,
            'stages': _renewal_workflow_stages(stage_obj.renewal) if stage_obj.workflow_type == 'renewal' else _main_workflow_stages(permit),
        })
    if permit.status == 'rejected':
        return render(request, 'work_permit/wp_email_action_done.html', {
            'message': 'This permit has already been rejected.',
            'permit': permit,
            'stages': _renewal_workflow_stages(stage_obj.renewal) if stage_obj.workflow_type == 'renewal' else _main_workflow_stages(permit),
        })

    if request.method == 'POST':
        chosen  = request.POST.get('action_choice', action)
        remarks = request.POST.get('remarks', '')
        stage_obj.status   = 'approved' if chosen == 'approve' else 'rejected'
        stage_obj.remarks  = remarks
        stage_obj.acted_at = timezone.now()
        # Resolve and store the approver so workflow can reference them
        approver = _find_wp_approver(
            stage_obj.approver_role,
            dept=stage_obj.approver_dept or None,
            designation=stage_obj.approver_desig or None,
        )
        if approver:
            stage_obj.approver = approver
        stage_obj.save()

        if chosen == 'reject':
            permit.status = 'rejected'
            permit.rejection_reason = remarks
            permit.save()
            _send_wp_status_email(permit, 'rejected')
        else:
            if not permit.actual_start:
                permit.actual_start = timezone.now()
                permit.save()
            _advance_wp_workflow(
                permit,
                request=request,
                renewal=stage_obj.renewal if stage_obj.workflow_type == 'renewal' else None,
            )

        return render(request, 'work_permit/wp_email_action_done.html', {
            'message': f'Stage {stage_obj.stage + 1} ({stage_obj.stage_label}) has been {stage_obj.status} successfully.',
            'permit': permit,
            'stages': _renewal_workflow_stages(stage_obj.renewal) if stage_obj.workflow_type == 'renewal' else _main_workflow_stages(permit),
        })

    return render(request, 'work_permit/wp_email_action.html', {
        'permit': permit, 'action': action, 'stage': stage_obj,
        'stages': _renewal_workflow_stages(stage_obj.renewal) if stage_obj.workflow_type == 'renewal' else _main_workflow_stages(permit),
    })


@login_required
def permit_renew(request, pk):
    """Initiator requests renewal — triggers 3-stage renewal workflow."""
    permit = get_object_or_404(WorkPermit, pk=pk)
    if permit.requested_by != request.user and not (request.user.is_superuser or request.user.role == 'administrator'):
        messages.error(request, 'Access denied.')
        return redirect('work_permit:detail', pk=pk)
    if not permit.renewal_required:
        messages.error(request, 'This permit does not have renewal enabled.')
        return redirect('work_permit:detail', pk=pk)
    if permit.status not in ('approved', 'closed'):
        messages.error(request, 'Only active or closed permits can be renewed.')
        return redirect('work_permit:detail', pk=pk)
    if request.method == 'POST':
        renew_comment = (request.POST.get('renew_comment') or '').strip()
        if not renew_comment:
            messages.error(request, 'Please add a renewal comment before submitting the renewal workflow.')
            return redirect('work_permit:detail', pk=pk)

        renewal_no = permit.renewals.count() + 1
        renewal = PermitRenewal.objects.create(
            permit=permit,
            renewal_no=renewal_no,
            requested_by=request.user,
            status='pending',
        )
        PermitComment.objects.create(
            permit=permit,
            author=request.user,
            comment=f'Renewal #{renewal_no} requested: {renew_comment}',
        )
        _create_wp_renewal_stages(permit, renewal)
        permit.status = 'pending'
        permit.actual_start = None
        permit.actual_end = None
        permit.closure_remarks = ''
        permit.save()
        _advance_wp_workflow(permit, request=request, renewal=renewal)
        messages.success(request, f'{permit.permit_number} renewal #{renewal_no} workflow started.')
    return redirect('work_permit:detail', pk=pk)


@login_required
def permit_reopen(request, pk):
    if not (request.user.is_superuser or request.user.role == 'administrator'):
        messages.error(request, 'Only administrators can re-open permits.')
        return redirect('work_permit:detail', pk=pk)
    permit = get_object_or_404(WorkPermit, pk=pk)
    if request.method == 'POST':
        permit.status = 'approved'
        permit.closure_remarks = ''
        permit.actual_end = None
        permit.save()
        messages.success(request, f'{permit.permit_number} re-opened.')
    return redirect('work_permit:detail', pk=pk)


@login_required
def permit_close(request, pk):
    permit = get_object_or_404(WorkPermit, pk=pk)
    if not (request.user.is_superuser or _can_approve(request.user) or permit.requested_by == request.user):
        messages.error(request, 'Access denied.')
        return redirect('work_permit:detail', pk=pk)

    if request.method == 'POST':
        form = PermitCloseForm(request.POST)
        if form.is_valid():
            permit.status = 'closed'
            permit.closure_remarks = form.cleaned_data['closure_remarks']
            permit.actual_end = timezone.now()
            permit.save()
            messages.success(request, f'{permit.permit_number} closed.')
            return redirect('work_permit:detail', pk=pk)
    else:
        form = PermitCloseForm()

    return render(request, 'work_permit/permit_close.html', {'permit': permit, 'form': form})


@login_required
def permit_print(request, pk):
    permit = get_object_or_404(WorkPermit, pk=pk)
    from accounts.models import SystemSetting
    setting = SystemSetting.get()
    main_stages = permit.approval_stages.filter(workflow_type='main').order_by('stage')
    # Group renewal stages by renewal record
    renewals_with_stages = []
    for renewal in permit.renewals.all().order_by('renewal_no'):
        stages = renewal.approval_stages.order_by('stage')
        if stages.exists():
            renewals_with_stages.append({'renewal': renewal, 'stages': stages})
    return render(request, 'work_permit/permit_print.html', {
        'permit': permit,
        'stages': main_stages,
        'renewals_with_stages': renewals_with_stages,
        'renewals': permit.renewals.all(),
        'fallback_stages': ['Issuer (HOD)', 'Safety Officer (HSEF)', 'Co-Permit (Third Party)'],
        'wp_doc_ref': getattr(setting, 'wp_doc_ref', 'PTW-F-001'),
        'wp_rev':     getattr(setting, 'wp_rev', '01'),
    })


@login_required
def approve_extension(request, ext_pk):
    ext = get_object_or_404(PermitExtension, pk=ext_pk)
    if not _can_approve(request.user):
        messages.error(request, 'Access denied.')
        return redirect('work_permit:detail', pk=ext.permit.pk)

    action = request.POST.get('action')
    if action == 'approve':
        ext.approved = True
        ext.approved_by = request.user
        ext.approved_at = timezone.now()
        ext.permit.end_datetime = ext.new_end_datetime
        ext.permit.save()
        ext.save()
        messages.success(request, 'Extension approved.')
    elif action == 'reject':
        ext.approved = False
        ext.approved_by = request.user
        ext.approved_at = timezone.now()
        ext.save()
        messages.warning(request, 'Extension rejected.')

    return redirect('work_permit:detail', pk=ext.permit.pk)


@login_required
def checklist_template_api(request):
    ptype = request.GET.get('type', '')
    return JsonResponse({'items': CHECKLIST_TEMPLATES.get(ptype, [])})


def _wp_filtered_qs(request):
    qs = WorkPermit.objects.select_related(
        'requested_by', 'hod_approved_by', 'safety_approved_by', 'final_approved_by'
    ).all()
    for field, param in [
        ('created_at__date__gte', 'date_from'),
        ('created_at__date__lte', 'date_to'),
        ('status',       'status'),
        ('permit_type',  'permit_type'),
        ('risk_level',   'risk_level'),
    ]:
        val = request.GET.get(param, '')
        if val:
            qs = qs.filter(**{field: val})
    return qs


@login_required
def permit_report(request):
    if not _can_view(request.user):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:index')

    view_mode = request.GET.get('view', 'detail')
    qs        = _wp_filtered_qs(request)
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    status_f  = request.GET.get('status', '')
    type_f    = request.GET.get('permit_type', '')
    risk_f    = request.GET.get('risk_level', '')

    from .models import PERMIT_TYPE_COLORS, PERMIT_TYPE_ICONS
    from django.db.models import Sum

    summary = {
        'total':     qs.count(),
        'active':    qs.filter(status='approved').count(),
        'pending':   qs.filter(status='pending').count(),
        'draft':     qs.filter(status='draft').count(),
        'closed':    qs.filter(status='closed').count(),
        'expired':   qs.filter(status='expired').count(),
        'rejected':  qs.filter(status='rejected').count(),
        'suspended': qs.filter(status='suspended').count(),
    }
    by_type = [
        {'code': code, 'label': label,
         'color': PERMIT_TYPE_COLORS.get(code, '#374151'),
         'icon':  PERMIT_TYPE_ICONS.get(code, 'bi-file-earmark-check'),
         'count': qs.filter(permit_type=code).count()}
        for code, label in PERMIT_TYPE_CHOICES
        if qs.filter(permit_type=code).exists()
    ]
    by_risk = {
        'critical': qs.filter(risk_level='critical').count(),
        'high':     qs.filter(risk_level='high').count(),
        'medium':   qs.filter(risk_level='medium').count(),
        'low':      qs.filter(risk_level='low').count(),
    }
    total_workers = qs.aggregate(total=Sum('workers_count'))['total'] or 0

    return render(request, 'work_permit/report.html', {
        'permits': qs, 'permit_type_choices': PERMIT_TYPE_CHOICES,
        'date_from': date_from, 'date_to': date_to,
        'status_f': status_f, 'type_f': type_f, 'risk_f': risk_f,
        'view_mode': view_mode, 'summary': summary,
        'by_type': by_type, 'by_risk': by_risk, 'total_workers': total_workers,
        'status_rows': [
            ('Active',    summary['active'],    'bg-success'),
            ('Pending',   summary['pending'],   'bg-warning text-dark'),
            ('Draft',     summary['draft'],     'bg-secondary'),
            ('Rejected',  summary['rejected'],  'bg-danger'),
            ('Suspended', summary['suspended'], 'bg-dark'),
            ('Closed',    summary['closed'],    'bg-secondary'),
            ('Expired',   summary['expired'],   'bg-danger'),
        ],
    })


@login_required
def report_export_excel(request):
    if not _can_view(request.user):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:index')

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from accounts.report_utils import add_excel_logo_and_note

    view_mode = request.GET.get('view', 'detail')
    qs        = _wp_filtered_qs(request)
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')

    wb  = openpyxl.Workbook()
    ws  = wb.active
    nf  = PatternFill('solid', fgColor='1e3a5f')
    hf  = Font(bold=True, color='FFFFFF')
    ctr = Alignment(horizontal='center')

    if view_mode == 'summary':
        ws.title = 'WP Summary Report'
        ws.merge_cells('A1:B1')
        ws['A1'] = 'WORK PERMIT — SUMMARY REPORT'
        ws['A1'].font = Font(bold=True, size=13, color='1e3a5f')
        ws['A1'].alignment = ctr
        ws.row_dimensions[1].height = 24
        ws.merge_cells('A2:B2')
        ws['A2'] = f'Generated: {timezone.localtime().strftime("%d %b %Y %H:%M")}  |  Period: {date_from or "All"} to {date_to or "All"}'
        ws['A2'].font = Font(italic=True, size=9, color='718096')
        ws['A2'].alignment = ctr
        ws.append([])

        from django.db.models import Sum
        sf = PatternFill('solid', fgColor='dc2626')
        bf = Font(bold=True, color='FFFFFF', size=9)
        sections = [
            ('BY STATUS', [
                ('Total Permits',  qs.count()),
                ('Active',         qs.filter(status='approved').count()),
                ('Pending',        qs.filter(status='pending').count()),
                ('Draft',          qs.filter(status='draft').count()),
                ('Closed',         qs.filter(status='closed').count()),
                ('Expired',        qs.filter(status='expired').count()),
                ('Rejected',       qs.filter(status='rejected').count()),
                ('Suspended',      qs.filter(status='suspended').count()),
                ('Total Workers',  qs.aggregate(t=Sum('workers_count'))['t'] or 0),
            ]),
            ('BY RISK LEVEL', [
                ('Critical', qs.filter(risk_level='critical').count()),
                ('High',     qs.filter(risk_level='high').count()),
                ('Medium',   qs.filter(risk_level='medium').count()),
                ('Low',      qs.filter(risk_level='low').count()),
            ]),
            ('BY PERMIT TYPE', [
                (label, qs.filter(permit_type=code).count())
                for code, label in PERMIT_TYPE_CHOICES
                if qs.filter(permit_type=code).exists()
            ]),
        ]
        for sec_title, rows in sections:
            ws.append([sec_title, 'COUNT'])
            for c in ws[ws.max_row]:
                c.fill = sf; c.font = bf; c.alignment = ctr
            for label, val in rows:
                ws.append([label, val])
                r = ws.max_row
                ws.cell(r, 1).font = Font(bold=True, size=9)
                ws.cell(r, 2).alignment = ctr
            ws.append([])
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 14
        add_excel_logo_and_note(ws, 2)
        fname = f'WP_Summary_Report_{timezone.localdate().strftime("%Y%m%d")}.xlsx'

    else:
        ws.title = 'WP Detail Report'
        headers = [
            '#', 'Permit No.', 'Permit Type', 'Title', 'Location', 'Plant Area',
            'Risk Level', 'Status', 'Shift', 'Workers',
            'Planned Start', 'Planned End', 'Actual Start', 'Actual End',
            'Requested By', 'Department', 'Contractor', 'Contractor Supervisor',
            'Hazards', 'Precautions', 'PPE Required',
            'Gas Test Required', 'Gas Test Result', 'Isolation Required',
            'HOD Approved By', 'HOD Approved At',
            'Safety Approved By', 'Safety Approved At',
            'Final Approved By', 'Final Approved At',
            'Rejection Reason', 'Closure Remarks', 'Created At',
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hf; cell.fill = nf; cell.alignment = ctr
            ws.column_dimensions[cell.column_letter].width = 20
        for i, p in enumerate(qs, 1):
            ws.append([
                i, p.permit_number, p.get_permit_type_display(),
                p.title, p.location, p.plant_area or '',
                p.get_risk_level_display(), p.get_status_display(),
                p.get_shift_display() if p.shift else '', p.workers_count,
                p.start_datetime.strftime('%d-%m-%Y %H:%M'),
                p.end_datetime.strftime('%d-%m-%Y %H:%M'),
                p.actual_start.strftime('%d-%m-%Y %H:%M') if p.actual_start else '',
                p.actual_end.strftime('%d-%m-%Y %H:%M') if p.actual_end else '',
                p.requested_by.employee_name, p.requested_by.department,
                p.contractor_name or '', p.contractor_supervisor or '',
                p.hazards or '', p.precautions or '', p.ppe_required or '',
                'Yes' if p.gas_test_required else 'No',
                p.gas_test_result or '',
                'Yes' if p.isolation_required else 'No',
                p.hod_approved_by.employee_name if p.hod_approved_by else '',
                p.hod_approved_at.strftime('%d-%m-%Y %H:%M') if p.hod_approved_at else '',
                p.safety_approved_by.employee_name if p.safety_approved_by else '',
                p.safety_approved_at.strftime('%d-%m-%Y %H:%M') if p.safety_approved_at else '',
                p.final_approved_by.employee_name if p.final_approved_by else '',
                p.final_approved_at.strftime('%d-%m-%Y %H:%M') if p.final_approved_at else '',
                p.rejection_reason or '', p.closure_remarks or '',
                p.created_at.strftime('%d-%m-%Y %H:%M'),
            ])
        add_excel_logo_and_note(ws, len(headers))
        fname = f'WP_Detail_Report_{timezone.localdate().strftime("%Y%m%d")}.xlsx'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


@login_required
def report_export_pdf(request):
    if not _can_view(request.user):
        messages.error(request, 'Access denied.')
        return redirect('dashboard:index')

    from reportlab.lib.pagesizes import A4, A3, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    from accounts.report_utils import build_pdf_header_table

    view_mode = request.GET.get('view', 'detail')
    qs        = _wp_filtered_qs(request)
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')

    navy  = colors.HexColor('#1e3a5f')
    alt   = colors.HexColor('#f8f9fa')
    WHITE = colors.white
    GREY  = colors.grey

    sub_s  = ParagraphStyle('s', fontName='Helvetica', fontSize=9,
                            alignment=TA_CENTER, textColor=GREY, spaceAfter=6)
    cell_s = ParagraphStyle('c', fontName='Helvetica', fontSize=6.5, leading=8)
    hdr_s  = ParagraphStyle('h', fontName='Helvetica-Bold', fontSize=6.5,
                            textColor=WHITE, alignment=TA_CENTER, leading=8)

    subtitle = (f'{"Summary" if view_mode == "summary" else "Detail"} Report'
                f' | Records: {qs.count()}'
                f' | Generated: {timezone.now().strftime("%d %b %Y %H:%M")}')
    if date_from or date_to:
        subtitle += f' | Period: {date_from or "All"} to {date_to or "All"}'

    buf = BytesIO()

    if view_mode == 'summary':
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                topMargin=12*mm, bottomMargin=10*mm,
                                leftMargin=10*mm, rightMargin=10*mm)
        elements = []
        elements.append(build_pdf_header_table('WORK PERMIT REPORT', subtitle, 277))
        elements.append(HRFlowable(width='100%', thickness=2, color=navy, spaceAfter=4*mm))

        from django.db.models import Sum
        total         = qs.count()
        total_workers = qs.aggregate(t=Sum('workers_count'))['t'] or 0

        sum_data = [['Metric', 'Count', 'Metric', 'Count']]
        sum_data += [
            ['Total Permits',  total,                                'Total Workers',  total_workers],
            ['Active',         qs.filter(status='approved').count(), 'Pending',        qs.filter(status='pending').count()],
            ['Draft',          qs.filter(status='draft').count(),    'Rejected',       qs.filter(status='rejected').count()],
            ['Closed',         qs.filter(status='closed').count(),   'Expired',        qs.filter(status='expired').count()],
            ['Critical Risk',  qs.filter(risk_level='critical').count(), 'High Risk',  qs.filter(risk_level='high').count()],
            ['Medium Risk',    qs.filter(risk_level='medium').count(),   'Low Risk',   qs.filter(risk_level='low').count()],
        ]
        sum_tbl = Table(sum_data, colWidths=[65*mm, 28*mm, 65*mm, 28*mm])
        sum_tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0),  navy),
            ('TEXTCOLOR',     (0,0), (-1,0),  WHITE),
            ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
            ('FONTNAME',      (0,1), (0,-1),  'Helvetica-Bold'),
            ('FONTNAME',      (2,1), (2,-1),  'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 8),
            ('ALIGN',         (1,0), (1,-1),  'CENTER'),
            ('ALIGN',         (3,0), (3,-1),  'CENTER'),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [WHITE, alt]),
            ('GRID',          (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
            ('TOPPADDING',    (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        elements.append(sum_tbl)
        elements.append(Spacer(1, 6*mm))

        type_data = [['Permit Type', 'Count']]
        for code, label in PERMIT_TYPE_CHOICES:
            cnt = qs.filter(permit_type=code).count()
            if cnt:
                type_data.append([label, cnt])
        if len(type_data) > 1:
            type_tbl = Table(type_data, colWidths=[110*mm, 30*mm])
            type_tbl.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (-1,0),  navy),
                ('TEXTCOLOR',     (0,0), (-1,0),  WHITE),
                ('FONTNAME',      (0,0), (-1,0),  'Helvetica-Bold'),
                ('FONTSIZE',      (0,0), (-1,-1), 8),
                ('ALIGN',         (1,0), (1,-1),  'CENTER'),
                ('ROWBACKGROUNDS',(0,1), (-1,-1), [WHITE, alt]),
                ('GRID',          (0,0), (-1,-1), 0.5, colors.HexColor('#dee2e6')),
                ('TOPPADDING',    (0,0), (-1,-1), 4),
                ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ]))
            elements.append(type_tbl)

    else:
        # A3 landscape — 420mm wide, all 16 columns fit perfectly
        doc = SimpleDocTemplate(buf, pagesize=landscape(A3),
                                topMargin=12*mm, bottomMargin=10*mm,
                                leftMargin=10*mm, rightMargin=10*mm)
        elements = []
        elements.append(build_pdf_header_table('WORK PERMIT DETAIL REPORT', subtitle, 410))
        elements.append(HRFlowable(width='100%', thickness=2, color=navy, spaceAfter=4*mm))

        col_headers = [
            '#', 'Permit No.', 'Permit Type', 'Title', 'Location',
            'Risk', 'Status', 'Workers', 'Start Date', 'End Date',
            'Requested By', 'Department', 'Contractor',
            'HOD Approved', 'Safety Approved', 'Final Approved',
        ]
        col_widths = [
            8*mm, 28*mm, 32*mm, 45*mm, 32*mm,
            16*mm, 18*mm, 14*mm, 22*mm, 22*mm,
            32*mm, 28*mm, 28*mm,
            22*mm, 22*mm, 22*mm,
        ]  # total = 391mm, fits in 410mm usable

        STATUS_C = {
            'approved': '#16a34a', 'pending': '#d97706', 'draft':     '#6b7280',
            'rejected': '#dc2626', 'closed':  '#374151', 'expired':   '#dc2626',
            'suspended':'#1f2937',
        }
        RISK_C = {'critical':'#dc2626','high':'#d97706','medium':'#2563eb','low':'#16a34a'}

        data = [[Paragraph(h, hdr_s) for h in col_headers]]
        for i, p in enumerate(qs, 1):
            hod_txt = (f'{p.hod_approved_by.employee_name}\n'
                       f'{p.hod_approved_at.strftime("%d %b %Y")}') if p.hod_approved_by else '-'
            saf_txt = (f'{p.safety_approved_by.employee_name}\n'
                       f'{p.safety_approved_at.strftime("%d %b %Y")}') if p.safety_approved_by else '-'
            fin_txt = (f'{p.final_approved_by.employee_name}\n'
                       f'{p.final_approved_at.strftime("%d %b %Y")}') if p.final_approved_by else '-'
            data.append([
                Paragraph(str(i), cell_s),
                Paragraph(f'<b>{p.permit_number}</b>', cell_s),
                Paragraph(p.get_permit_type_display(), cell_s),
                Paragraph(p.title, cell_s),
                Paragraph(p.location, cell_s),
                Paragraph(f'<font color="{RISK_C.get(p.risk_level,"#374151")}"><b>{p.get_risk_level_display()}</b></font>', cell_s),
                Paragraph(f'<font color="{STATUS_C.get(p.status,"#374151")}"><b>{p.get_status_display()}</b></font>', cell_s),
                Paragraph(str(p.workers_count), cell_s),
                Paragraph(p.start_datetime.strftime('%d %b %Y'), cell_s),
                Paragraph(p.end_datetime.strftime('%d %b %Y'), cell_s),
                Paragraph(p.requested_by.employee_name, cell_s),
                Paragraph(p.requested_by.department, cell_s),
                Paragraph(p.contractor_name or '-', cell_s),
                Paragraph(hod_txt, cell_s),
                Paragraph(saf_txt, cell_s),
                Paragraph(fin_txt, cell_s),
            ])

        if len(data) == 1:
            data.append([Paragraph('No records found.', cell_s)] + [Paragraph('', cell_s)] * 15)

        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0),  navy),
            ('ROWBACKGROUNDS',(0,1), (-1,-1), [WHITE, alt]),
            ('GRID',          (0,0), (-1,-1), 0.4, colors.HexColor('#dee2e6')),
            ('LINEABOVE',     (0,0), (-1,0),  1.5, navy),
            ('LINEBELOW',     (0,0), (-1,0),  1.5, navy),
            ('FONTSIZE',      (0,0), (-1,-1), 6.5),
            ('TOPPADDING',    (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN',         (0,0), (0,-1),  'CENTER'),
            ('ALIGN',         (7,0), (7,-1),  'CENTER'),
            ('ALIGN',         (13,0),(15,-1), 'CENTER'),
        ]))
        elements.append(tbl)

    elements.append(Spacer(1, 6*mm))
    elements.append(Paragraph('This is a computer generated report. — Unity Cement ERP System', sub_s))
    doc.build(elements)
    buf.seek(0)
    fname = f'WP_{"Summary" if view_mode == "summary" else "Detail"}_Report_{timezone.localdate().strftime("%Y%m%d")}.pdf'
    response = HttpResponse(buf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response
